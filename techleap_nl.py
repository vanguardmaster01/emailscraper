import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import json
import pandas as pd
from openpyxl import Workbook
import random
import requests

def get_status(url):
    r = requests.get(url)
    soupGet = BeautifulSoup(r.text, "html.parser")
    try:
        website= soupGet.find('a', {'class': "item-details-info__url"}).text.strip()
    except:
        website = ''

    try:
        company_name=soupGet.find('h1', {'class': "name"}).text.strip()
    except:
        company_name = ''

    try:
        location=soupGet.find('div', {'itemprop': "location"}).text.strip()
    except:
        location=''

    return company_name ,website,location

def login():
    firefox_options = Options()
    # firefox_options.add_argument("--headless")
    driver = webdriver.Firefox(firefox_options=firefox_options)
    driver.set_window_size(1920, 1080)
    url = "https://finder.techleap.nl/companies.startups/f/all_locations/allof_Netherlands?row_index=0"
    driver.get(url)
    time.sleep(random.randint(15, 30))
    driver.find_element_by_class_name('login-button').click()
    time.sleep(15)
    try:
        driver.find_element_by_id('login-form__first-input').send_keys('hta@siboard.com')
    except:
        time.sleep(10)
        driver.find_element_by_id('login-form__first-input').send_keys('hta@siboard.com')
    driver.find_element_by_id('password').send_keys('Q8p6854944123!')

    driver.find_elements_by_class_name('login-button')[1].click()
    time.sleep(random.randint(4, 10))
    return driver

def techleap():
    try:
        wb = Workbook()
        worksheet1 = wb.active
        worksheet1.cell(1, 1, 'Company_name')
        worksheet1.cell(1, 2, 'website')
        worksheet1.cell(1, 3, 'location')
        recordpos = 2

        driver=login()
        try:
            df=[]
            pageNum=0
            for row in range(1500):
                try:
                    if pageNum!=0:
                        driver.get('https://finder.techleap.nl/companies.startups/f/all_locations/allof_Netherlands?row_index=' + str(pageNum))
                        time.sleep(random.randint(10, 20))
                    else:
                        time.sleep(random.randint(5, 20))

                    pageNum=pageNum+25
                    listOfUrlsElement=driver.find_elements_by_xpath("//div[@class='type-element type-element--h3 hbox entity-name__name entity-name__name--black']")
                    listofUrls=[]
                    for item in listOfUrlsElement:
                        itemUrl=item.find_element_by_xpath("a").get_attribute('href')
                        if itemUrl in df:
                            continue
                        else:
                            listofUrls.append(itemUrl)
                            df.append(itemUrl)

                    for item in listofUrls:
                        company_name ,website,location= get_status(item)
                        worksheet1.cell(recordpos, 1, company_name)
                        worksheet1.cell(recordpos, 2, website)
                        worksheet1.cell(recordpos, 3, location)
                        wb.save('techleap.xlsx')
                        recordpos = recordpos + 1
                        print(str(recordpos - 2) + ' )Done :' + str(company_name))
                except Exception as e:
                    print(str(e))
                    pass

        except Exception as e:
            print(str(e))
            driver.close()
            driver.quit()
        driver.close()
        driver.quit()
        read_file = pd.read_excel(r'techleap.xlsx')
        read_file.to_csv(r'techleap.csv', index=None, header=True)

    except Exception as e:
        print(str(e))


if __name__ == "__main__":
    techleap()
